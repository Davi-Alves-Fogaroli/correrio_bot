from dotenv import load_dotenv
import os
import time
import logging
# Error libs
from selenium.common.exceptions import WebDriverException
# selenium
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
# excel lib
import openpyxl

load_dotenv()

# op = Options 
# AttributeError: type object 'Options' has no attribute '_ignore_local_proxy'
#opitions.headless = True <- Hability to run in server"

try:
    # configure and define webdriver
    service = Service(os.getenv("DRIVER_PATH"))
    driver = webdriver.Chrome(service=service)
    driver.get("https://buscacepinter.correios.com.br/app/endereco/index.php")
    driver.maximize_window()


except WebDriverException :
    print("\n\n\n","ERROR: Verify you crhomedriver driver version and yor connection","\n\n\n")

logging.info("Starting")

ceps_no_address = openpyxl.load_workbook("excel\excel_desatualizado.xlsx")

logging.info("Opening ceps file")
planilha1 = ceps_no_address["planilha 1"]
ceps = [row[0].value for row in planilha1.iter_rows(min_row=2, max_row=11)]

def get_addresses(driver,ceps):
    address_list = []
    temp_address_list=[]

    for cep in ceps:
        logging.info(f"Getting address for cep: {cep}")
        driver.find_element(By.ID, "endereco").send_keys(f"{cep}" + Keys.ENTER)
        
        time.sleep(1) 
        td = driver.find_elements(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td')
        
        temp_address_list=[]
        counter=1
        for x in td:
            if counter == 1:
                temp_address_list.append(x.text)

            elif counter == 2:
                temp_address_list.append(x.text)

            elif counter == 3:
                temp_address_list.append(x.text[:-3])
                temp_address_list.append(x.text[-2:])

            elif counter == 4:
                temp_address_list.append(x.text)
                address_list.append(temp_address_list)
            counter+=1

        WebDriverWait(driver, timeout=10).until(EC.element_to_be_clickable((By.ID, "btn_nbusca"))).click()

    return address_list
logging.info("Getting addresses")

addresses = get_addresses(driver,ceps)
logging.info("Saving addresses")

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "planilha 1"
sheet.append(["CEP", "Logradouro", "Bairro", "Cidade", "Estado"])

for address in addresses:
    logging.info(f"Saving address: {address}")
    sheet.append(address)

wb.save("excel/excel_atualizado.xlsx")
logging.info("Finished")
