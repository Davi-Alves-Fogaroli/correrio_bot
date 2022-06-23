import time
from pycep_correios import get_address_from_cep, WebService
import openpyxl
import logging
import os

logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s:%(levelname)s:%(message)s',
                    handlers=[
                        logging.FileHandler("debug.log"),
                        logging.StreamHandler()  # print to console
                    ])

logging.info("Starting")

ceps_file = os.path.join(os.path.dirname(__file__), "excel/excel_desatualizado.xlsx")
output_file = os.path.join(os.path.dirname(__file__), "excel/excel_atualizado.xlsx")

logging.info("Opening ceps file")

zip_code_without_address = openpyxl.load_workbook(ceps_file)
planilha1 = zip_code_without_address["planilha 1"]
ceps = [row[0].value for row in planilha1.iter_rows(min_row=2, max_row=11)]

def get_addresses(ceps: list) -> list:
    address_list = []

    for cep in ceps:
        logging.info(f"Getting address for cep: {cep}")
        
        try:
            address = get_address_from_cep(cep, webservice=WebService.CORREIOS)
            logging.info(f"Address: {address}")
            address_list.append(address)
            time.sleep(2)
        
        except Exception as e:
            logging.error(f"Getting {e} for cep: {cep}")
            time.sleep(2)
            continue

    return address_list

logging.info("Getting addresses")

addresses = get_addresses(ceps)

logging.info("Saving addresses")

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "planilha 1"
sheet.append(["CEP", "Logradouro", "Bairro", "Cidade", "Estado"])

for address in addresses:
    logging.info(f"Saving address: {address}")
    sheet.append([address["cep"], address["logradouro"],
                 address["bairro"], address["cidade"], address["uf"]])

wb.save(output_file)

logging.info("Finished")
